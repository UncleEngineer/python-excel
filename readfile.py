from openpyxl import load_workbook
from openpyxl import Workbook

excelfile = Workbook()
sheet = excelfile.active

readfile = load_workbook('name.xlsx')
current_sheet = readfile.active

#print(current_sheet['A1'].value)

allname = []

for i in range(1,8):
    allname.append(current_sheet.cell(row=i,column=1).value)

print(allname)


checkname = {} #นับว่าแต่ละตัวอักษรมีเท่าไร

for nm in allname:
    #print(nm[0])

    if nm[0].upper() not in checkname.keys():
        checkname[nm[0].upper()] = 1
        sheet[nm[0].upper()+'1'] =  nm
    else:
        checkname[nm[0].upper()] = checkname[nm[0].upper()] + 1
        sheet[nm[0].upper() + str(checkname[nm[0].upper()])] = nm

print(checkname)

excelfile.save('Allname2.xlsx')
